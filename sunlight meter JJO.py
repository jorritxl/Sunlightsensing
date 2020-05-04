#helaas is het originelen arduino file niet meer te vinden het programma staat op de artduino mini zelf alleen en werkt prima hiermee. het file samenvoegsel komt wel het dichtsts bij. Denk dat ie werkt maar durf het niet. 2 fotos van de opstelling staan in jorrits dropbox. de output naar de serial monitor moet dit zijn, hum,temp,brit,light de rest doet python zelf

import serial # import Serial Library
import numpy  # Import numpy
import matplotlib.pyplot as plt #import matplotlib library
from drawnow import *
import datetime
import time
import csv
from openpyxl import load_workbook

tempAA= []
tempBB= []
britnesss= []
tijdd= []

wb = load_workbook('dekn.xlsx')
ws = wb.active

#--------
now = datetime.datetime.now()
#--------
arduinoData = serial.Serial('com10', 9600) #Creating our serial object named arduinoData
plt.ion()
cnt=0

row = ['String', ' String']

with open('sunlightmeter1.csv', 'r') as readFile:
    reader = csv.reader(readFile)
    lines = list(reader)
    lines.insert(1, row)

##cnt=0

def makeFig():
    plt.ylim(0,100)                                 #Set y min and max values
    plt.title('Suntime meter JJO')                  #Plot the title
    plt.grid(True)                                  #Turn the grid on
    plt.ylabel('lux')
    plt.xlabel('tijd')
    plt.plot(tijdd, britnesss, label='britness',color='#0066FF')       #plot the temperature
    plt.legend(loc='upper left') 
    plt2=plt.twinx()
    plt.ylim(15,25)
    plt2.plot(tijdd, tempBB, label='tempB',color='#ff9900')       #plot the temperature
    plt2.set_ylabel('lux (*C)')
    plt2.legend(loc='upper right')                  #plot the legend
    
while True: # While loop that loops forever
    while (arduinoData.inWaiting()==0): #Wait here until there is data
        pass #do nothing
    timestamp = str(datetime.datetime.fromtimestamp(time.time()).strftime('%H:%M:%S'))
    line = arduinoData.readline()          #read the line of text from the serial port
    leknn = str(line, 'utf-8')              #byte to str
    lekn = leknn +"," + timestamp
    
    dataArray = lekn.split(',')            #Split it into an array called dataArray
    tempA = float( dataArray[0])            #Convert first element to floating number and put in temp
    tempB =    float( dataArray[1])            #Convert second element to floating number and put in P
    britness =    float( dataArray[2])
    tijd = ( dataArray[3])
    print(tempA)
    print(tempB)
    print(britness)
    print(tijd)
    tempAA.append(tempA)                     #Build our tempF array by appending temp readings
    tempBB.append(tempB)
    britnesss.append(britness)                   #Building our pressure array by appending P readings
    tijdd.append(tijd)
    drawnow(makeFig)                       #Call drawnow to update our live graph
    plt.pause(.000001)                     #Pause Briefly. Important to keep drawnow from crashing
    cnt=cnt+1
    print(cnt)

    ws.cell(row=(cnt)+4, column=1, value=tijd)
    ws.cell(row=(cnt)+4, column=2, value=tempA)
    ws.cell(row=(cnt)+4, column=3, value=tempB)
    ws.cell(row=(cnt)+4, column=4, value=britness)
    
    wb.save("dekn.xlsx")


    
##    cnt=cnt+1
##    if(cnt>20):                            #If you have 50 or more points, delete the first one from the array
##        tempAA.pop(0)                       #This allows us to just see the last 50 data points
    with open('sunlightmeter1.csv', 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=' ', quoting=csv.QUOTE_MINIMAL)
            spamwriter.writerow([tempA]+[","]+[tempB]+[","]+[britness]+[","]+[tijd])


